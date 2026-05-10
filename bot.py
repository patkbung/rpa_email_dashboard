"""
What it does
────────────
1.  Creates all required folders if they are missing.
2.  Creates reports/daily_report.xlsx with a styled header if it does not exist.
3.  Generates sample test files in test_files/ and organises them (local demo).
4.  Connects to Gmail via IMAP using credentials from .env.
5.  Searches both INBOX and [Gmail]/Spam for emails whose subject contains EMAIL_SEARCH_KEYWORD.
6.  Downloads every attachment into temp_downloads/.
7.  Passes each downloaded file to move_file_to_category() → downloads/<type>/.
8.  Logs every result to reports/daily_report.xlsx via write_log().
"""

import os
import imaplib
import email
import shutil
from email.header import decode_header
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

from config import (
    BASE_DIR,
    DOWNLOADS_DIR,
    REPORTS_DIR,
    TEST_FILES_DIR,
    REPORT_FILE,
    CATEGORY_DIRS,
    EXTENSION_MAP,
    REPORT_COLUMNS,
)

# ── Load .env ─────────────────────────────────────────────────────────────────
load_dotenv()

EMAIL_ADDRESS        = os.getenv("EMAIL_ADDRESS", "").strip()
EMAIL_APP_PASSWORD   = os.getenv("EMAIL_APP_PASSWORD", "").strip()
EMAIL_SEARCH_KEYWORD = os.getenv("EMAIL_SEARCH_KEYWORD", "").strip()

# Temporary folder for raw attachment downloads before categorisation
TEMP_DOWNLOADS_DIR = os.path.join(BASE_DIR, "temp_downloads")

IMAP_HOST = "imap.gmail.com"
IMAP_PORT = 993


# ══════════════════════════════════════════════════════════════════════════════
# 1. FOLDER SETUP
# ══════════════════════════════════════════════════════════════════════════════

def create_folders() -> None:
    """Create all required directories (downloads sub-folders, reports, test_files, temp_downloads)."""
    folders = (
        [DOWNLOADS_DIR, REPORTS_DIR, TEST_FILES_DIR, TEMP_DOWNLOADS_DIR]
        + list(CATEGORY_DIRS.values())
    )
    for folder in folders:
        os.makedirs(folder, exist_ok=True)
    print("[✔] All folders verified / created.")


# ══════════════════════════════════════════════════════════════════════════════
# 2. EXCEL REPORT SETUP
# ══════════════════════════════════════════════════════════════════════════════

def _style_header_row(ws) -> None:
    """Apply formatting to the header row of the worksheet."""
    header_fill  = PatternFill("solid", fgColor="1F4E79")   # dark blue
    accent_fill  = PatternFill("solid", fgColor="2E75B6")   # medium blue
    header_font  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
        cell.font      = header_font
        cell.fill      = header_fill if col_idx % 2 == 1 else accent_fill
        cell.alignment = center_align
        cell.border    = thin_border

    ws.row_dimensions[1].height = 22


def _set_column_widths(ws) -> None:
    """Set sensible column widths."""
    widths = {
        "date": 12, "time": 10, "email_subject": 35,
        "sender": 28, "file_name": 30, "file_type": 10,
        "save_path": 45, "status": 12, "error_message": 35,
    }
    for col_idx, col_name in enumerate(REPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 15)


def create_report_if_missing() -> None:
    """Create reports/daily_report.xlsx with a styled header if the file does not exist."""
    if os.path.exists(REPORT_FILE):
        print(f"[✔] Report already exists: {REPORT_FILE}")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None, "openpyxl returned no active worksheet"
    ws.title = "Daily Report"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(REPORT_COLUMNS))}1"

    _style_header_row(ws)
    _set_column_widths(ws)

    wb.save(REPORT_FILE)
    print(f"[✔] Report created: {REPORT_FILE}")


# ══════════════════════════════════════════════════════════════════════════════
# 3. CORE FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def get_file_type(file_name: str) -> str:
    """
    Return the category string for a given file name based on its extension.

    Parameters
    ----------
    file_name : str
        Name (or path) of the file.

    Returns
    -------
    str
        One of: 'pdf', 'word', 'excel', 'image', 'others'.
    """
    _, ext = os.path.splitext(file_name.lower())
    return EXTENSION_MAP.get(ext, "others")


def write_log(
    email_subject: str,
    sender: str,
    file_name: str,
    file_type: str,
    save_path: str,
    status: str,
    error_message: str = "",
) -> None:
    """
    Append one styled row to reports/daily_report.xlsx.

    Parameters
    ----------
    email_subject  : str   Subject line of the source email (or 'N/A').
    sender         : str   Sender address (or 'N/A').
    file_name      : str   Original file name.
    file_type      : str   Category ('pdf', 'word', …).
    save_path      : str   Absolute destination path.
    status         : str   'SUCCESS' or 'FAILED'.
    error_message  : str   Details if status is 'FAILED'.
    """
    now  = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    wb = openpyxl.load_workbook(REPORT_FILE)
    ws = wb.active
    assert ws is not None, "openpyxl returned no active worksheet"

    even_fill   = PatternFill("solid", fgColor="DCE6F1")
    odd_fill    = PatternFill("solid", fgColor="FFFFFF")
    ok_font     = Font(color="1F6B35", bold=True, name="Calibri", size=10)
    err_font    = Font(color="C00000", bold=True, name="Calibri", size=10)
    base_font   = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    row_data = [date, time, email_subject, sender, file_name,
                file_type, save_path, status, error_message]

    next_row = ws.max_row + 1
    row_fill = even_fill if next_row % 2 == 0 else odd_fill

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.border = thin_border
        cell.fill   = row_fill

        col_name = REPORT_COLUMNS[col_idx - 1]
        if col_name == "status":
            cell.font      = ok_font if status == "SUCCESS" else err_font
            cell.alignment = center_align
        elif col_name in ("date", "time", "file_type"):
            cell.font      = base_font
            cell.alignment = center_align
        else:
            cell.font      = base_font
            cell.alignment = left_align

    wb.save(REPORT_FILE)


def move_file_to_category(
    src_path: str,
    email_subject: str = "N/A",
    sender: str = "N/A",
) -> dict:
    """
    Move *src_path* to the appropriate downloads/<category>/ folder and log the result.

    Parameters
    ----------
    src_path      : str   Absolute path to the source file.
    email_subject : str   Metadata to store in the log (default 'N/A').
    sender        : str   Metadata to store in the log (default 'N/A').

    Returns
    -------
    dict  {'file': str, 'category': str, 'dest': str, 'status': str}
    """
    file_name = os.path.basename(src_path)
    category  = get_file_type(file_name)
    dest_dir  = CATEGORY_DIRS[category]
    dest_path = os.path.join(dest_dir, file_name)

    # Avoid silently overwriting duplicates
    if os.path.exists(dest_path):
        stem, ext = os.path.splitext(file_name)
        suffix    = datetime.now().strftime("%Y%m%d%H%M%S%f")
        dest_path = os.path.join(dest_dir, f"{stem}_{suffix}{ext}")

    try:
        shutil.move(src_path, dest_path)
        write_log(
            email_subject=email_subject,
            sender=sender,
            file_name=file_name,
            file_type=category,
            save_path=dest_path,
            status="SUCCESS",
        )
        print(f"  [✔] {file_name:40s} → downloads/{category}/")
        return {"file": file_name, "category": category, "dest": dest_path, "status": "SUCCESS"}

    except Exception as exc:
        err = str(exc)
        write_log(
            email_subject=email_subject,
            sender=sender,
            file_name=file_name,
            file_type=category,
            save_path=dest_path,
            status="FAILED",
            error_message=err,
        )
        print(f"  [✘] {file_name:40s} ERROR: {err}")
        return {"file": file_name, "category": category, "dest": dest_path,
                "status": "FAILED", "error": err}


# ══════════════════════════════════════════════════════════════════════════════
# 4. GMAIL IMAP — HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _decode_mime_words(raw: str) -> str:
    """Decode RFC-2047 encoded header words into a plain Unicode string."""
    parts = decode_header(raw or "")
    decoded = []
    for part, charset in parts:
        if isinstance(part, bytes):
            decoded.append(part.decode(charset or "utf-8", errors="replace"))
        else:
            decoded.append(part)
    return "".join(decoded).strip()


def _safe_filename(name: str) -> str:
    """Strip path separators and null bytes from an attachment filename."""
    name = os.path.basename(name)
    keepchars = (" ", ".", "_", "-")
    return "".join(c for c in name if c.isalnum() or c in keepchars).strip() or "attachment"


def _connect_imap() -> imaplib.IMAP4_SSL:
    """
    Open an authenticated IMAP4_SSL connection to Gmail.

    Returns
    -------
    imaplib.IMAP4_SSL
        Authenticated IMAP connection. No mailbox is pre-selected;
        the caller must call mail.select() before searching.

    Raises
    ------
    EnvironmentError
        If EMAIL_ADDRESS or EMAIL_APP_PASSWORD are missing from .env.
    imaplib.IMAP4.error
        On authentication failure.
    """
    if not EMAIL_ADDRESS or not EMAIL_APP_PASSWORD:
        raise EnvironmentError(
            "EMAIL_ADDRESS and EMAIL_APP_PASSWORD must be set in .env\n"
            "Generate an App Password at: https://myaccount.google.com/apppasswords"
        )

    print(f"[→] Connecting to {IMAP_HOST} as {EMAIL_ADDRESS} …")
    mail = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT)

    try:
        mail.login(EMAIL_ADDRESS, EMAIL_APP_PASSWORD)
    except imaplib.IMAP4.error as exc:
        raise imaplib.IMAP4.error(
            f"Login failed for {EMAIL_ADDRESS}. "
            "Make sure you are using a Gmail App Password, not your regular password. "
            f"Details: {exc}"
        ) from exc

    print("[✔] Authenticated successfully.")
    return mail


# ══════════════════════════════════════════════════════════════════════════════
# 5. GMAIL IMAP — MAIN DOWNLOAD FUNCTION
# ══════════════════════════════════════════════════════════════════════════════

def _process_mailbox(
    mail: imaplib.IMAP4_SSL,
    mailbox: str,
    search_criterion: str,
) -> int:
    """
    Select *mailbox*, search using *search_criterion*, download all attachments
    from matching emails, and pass each file to move_file_to_category().

    Parameters
    ----------
    mail             : authenticated IMAP4_SSL connection
    mailbox          : e.g. 'INBOX' or '[Gmail]/Spam'
    search_criterion : IMAP search string, e.g. '(SUBJECT "Invoice")'

    Returns
    -------
    int   Number of attachments successfully downloaded and moved.
    """
    # ── Select mailbox (quote the name to handle special chars like [ ] /) ───
    # IMAP requires mailbox names with special characters to be double-quoted.
    quoted_mailbox = f'"{mailbox}"' if any(c in mailbox for c in ' []()') else mailbox
    try:
        status, data = mail.select(quoted_mailbox, readonly=True)
    except imaplib.IMAP4.error as exc:
        print(f"  [\u2718] Cannot open mailbox '{mailbox}': {exc} \u2014 skipping.")
        return 0

    if status != "OK":
        msg = data[0].decode() if data and data[0] else "unknown error"
        print(f"  [\u26a0] Mailbox '{mailbox}' returned status '{status}' ({msg}) \u2014 skipping.")
        return 0

    print(f"  [✔] Mailbox '{mailbox}' opened.")

    # ── Search ────────────────────────────────────────────────────────────────
    try:
        status, data = mail.search(None, search_criterion)
    except imaplib.IMAP4.error as exc:
        print(f"  [✘] SEARCH failed in '{mailbox}': {exc}")
        return 0

    if status != "OK" or not data or not data[0]:
        print(f"  [ℹ] No matching emails in '{mailbox}'.")
        return 0

    email_ids  = data[0].split()
    total_msgs = len(email_ids)
    print(f"  [→] {total_msgs} email(s) found in '{mailbox}'.\n")

    total_downloaded = 0

    # ── Iterate emails ────────────────────────────────────────────────────────
    for idx, eid in enumerate(email_ids, start=1):
        try:
            status, msg_data = mail.fetch(eid, "(RFC822)")
        except imaplib.IMAP4.error as exc:
            print(f"    [✘] Could not fetch email ID {eid.decode()} in '{mailbox}': {exc}")
            continue

        if status != "OK" or not msg_data:
            print(f"    [⚠] Empty response for email ID {eid.decode()} — skipping.")
            continue

        first = msg_data[0]
        if not isinstance(first, tuple):
            print(f"    [⚠] Unexpected fetch response for email ID {eid.decode()} — skipping.")
            continue
        raw_email = first[1]
        if not isinstance(raw_email, bytes):
            print(f"    [⚠] Non-bytes payload for email ID {eid.decode()} — skipping.")
            continue
        msg = email.message_from_bytes(raw_email)

        subject = _decode_mime_words(msg.get("Subject", "(no subject)"))
        sender  = _decode_mime_words(msg.get("From",    "(unknown sender)"))

        print(f"    [{idx}/{total_msgs}] Mailbox : {mailbox}")
        print(f"              From    : {sender}")
        print(f"              Subject : {subject}")

        has_attachment = False

        # ── Walk MIME parts ───────────────────────────────────────────────────
        for part in msg.walk():
            if part.get_content_maintype() == "multipart":
                continue
            if part.get("Content-Disposition") is None:
                continue

            raw_filename = part.get_filename()
            if not raw_filename:
                continue

            file_name = _safe_filename(_decode_mime_words(raw_filename))
            temp_path = os.path.join(TEMP_DOWNLOADS_DIR, file_name)

            # Avoid overwriting temp files with the same name
            if os.path.exists(temp_path):
                stem, ext = os.path.splitext(file_name)
                suffix    = datetime.now().strftime("%Y%m%d%H%M%S%f")
                file_name = f"{stem}_{suffix}{ext}"
                temp_path = os.path.join(TEMP_DOWNLOADS_DIR, file_name)

            # ── Save to temp_downloads/ ───────────────────────────────────────
            try:
                payload = part.get_payload(decode=True)
                if not isinstance(payload, bytes):
                    print(f"      [⚠] Empty or non-bytes payload for {file_name} — skipping.")
                    continue

                with open(temp_path, "wb") as fh:
                    fh.write(payload)

                print(f"      [↓] Downloaded : {file_name}  (mailbox: {mailbox})")
                has_attachment = True

            except OSError as exc:
                err = str(exc)
                print(f"      [✘] Could not save {file_name}: {err}")
                write_log(
                    email_subject=subject,
                    sender=sender,
                    file_name=file_name,
                    file_type=get_file_type(file_name),
                    save_path=temp_path,
                    status="FAILED",
                    error_message=f"[{mailbox}] Download error: {err}",
                )
                continue

            # ── Move to category folder ───────────────────────────────────────
            result = move_file_to_category(temp_path, email_subject=subject, sender=sender)
            if result["status"] == "SUCCESS":
                total_downloaded += 1

        if not has_attachment:
            print("      [ℹ] No attachments in this email.")

        print()  # blank line between emails

    return total_downloaded

def _find_spam_folder(mail: imaplib.IMAP4_SSL) -> str:
    """
    Detect the real spam/junk folder name by listing all IMAP mailboxes.

    Strategy (most-reliable first):
      1. Match the \\Junk IMAP attribute flag. Gmail always sets this on the
         spam folder regardless of the account language. Works even when the
         folder name is encoded in modified UTF-7 (Thai, Chinese, etc.).
      2. Decode modified UTF-7 names and match known spam keywords as fallback.

    Returns the raw folder name string (as returned by the server).
    Falls back to '[Gmail]/Spam' as a last resort.
    """
    import re as _re, base64 as _b64

    def _decode_mutf7(encoded: str) -> str:
        """Decode IMAP modified UTF-7 to plain Unicode."""
        def _token(m):
            b64 = m.group(1).replace(",", "/")
            if not b64:
                return "&"
            try:
                pad = b64 + "=" * (-len(b64) % 4)
                return _b64.b64decode(pad).decode("utf-16-be", errors="replace")
            except Exception:
                return m.group(0)
        return _re.sub(r"&([^-]*)-", _token, encoded)

    spam_keywords = [
        "spam", "junk",
        "\u0e2a\u0e41\u0e1b\u0e21",   # Thai: \u0e2a\u0e41\u0e1b\u0e21
        "\u5783\u573e\u90ae\u4ef6",   # Chinese (simplified)
        "\u5783\u573e\u90f5\u4ef6",   # Chinese (traditional)
        "correo no deseado",          # Spanish
        "ind\u00e9sirables",          # French
        "spam-ordner",                # German
    ]

    try:
        status, mailbox_list = mail.list()
    except imaplib.IMAP4.error as exc:
        print(f"  [\u26a0] Could not list mailboxes: {exc}")
        return ""

    if status != "OK" or not mailbox_list:
        return ""

    keyword_match = ""

    for item in mailbox_list:
        if not item:
            continue
        # Normalise to str — items can be bytes, str, or tuple
        if isinstance(item, bytes):
            line: str = item.decode("utf-8", errors="replace")
        elif isinstance(item, str):
            line = item
        else:
            # Unexpected type (e.g. tuple from some servers) — skip
            continue

        # Extract raw folder name from the last quoted segment
        m = _re.search(r"\'([^\']+)\'\s*$", line)
        if not m:
            m = _re.search(r'"([^"]+)"\s*$', line)
        folder_raw = m.group(1) if m else line.rsplit(None, 1)[-1].strip('"')

        # Strategy 1: \Junk IMAP attribute (language-independent)
        if "\\Junk" in line:
            print(f"  [\u2714] Spam folder detected (\\Junk flag): {folder_raw}")
            return folder_raw

        # Strategy 2: keyword match after UTF-7 decode
        folder_unicode = _decode_mutf7(folder_raw).lower()
        if any(kw in folder_unicode for kw in spam_keywords):
            readable = _decode_mutf7(folder_raw)
            print(f"  [\u2714] Spam folder detected (keyword): {folder_raw} -> \'{readable}\'")
            keyword_match = folder_raw

    if keyword_match:
        return keyword_match

    print("  [\u2139] Could not auto-detect spam folder — will try '[Gmail]/Spam' as fallback.")
    return "[Gmail]/Spam"

def download_gmail_attachments() -> int:
    """
    Connect to Gmail and search both INBOX and [Gmail]/Spam for emails whose
    subject matches EMAIL_SEARCH_KEYWORD.  Downloads all attachments to
    temp_downloads/, then passes each file to move_file_to_category().

    Mailboxes that cannot be opened are skipped with a clear message.

    Returns
    -------
    int   Total attachments successfully downloaded and moved across all mailboxes.
    """
    if not EMAIL_SEARCH_KEYWORD:
        print("[⚠] EMAIL_SEARCH_KEYWORD is empty in .env — skipping Gmail step.")
        return 0

    # ── Connect (authenticate only — no mailbox selected yet) ────────────────
    try:
        mail = _connect_imap()
    except EnvironmentError as exc:
        print(f"\n[✘] Configuration error:\n    {exc}\n")
        return 0
    except imaplib.IMAP4.error as exc:
        print(f"\n[✘] IMAP authentication error:\n    {exc}\n")
        return 0
    except OSError as exc:
        print(f"\n[✘] Network error connecting to {IMAP_HOST}:\n    {exc}\n")
        return 0

    # ── Detect spam folder name (varies by Gmail language) ───────────────────
    print("[→] Detecting spam folder name …")
    spam_folder = _find_spam_folder(mail)

    mailboxes        = ["INBOX"] + ([spam_folder] if spam_folder else [])
    search_criterion = f'(SUBJECT "{EMAIL_SEARCH_KEYWORD}")'
    total_downloaded = 0

    for mailbox in mailboxes:
        print(f"\n[→] Searching mailbox: {mailbox} …")
        total_downloaded += _process_mailbox(mail, mailbox, search_criterion)

    # ── Logout ────────────────────────────────────────────────────────────────
    try:
        mail.logout()
        print("\n[✔] IMAP connection closed.")
    except Exception:
        pass  # best-effort logout

    return total_downloaded


# ══════════════════════════════════════════════════════════════════════════════
# 6. TEST-FILE GENERATION  (unchanged from Phase 1)
# ══════════════════════════════════════════════════════════════════════════════

def create_sample_test_files() -> None:
    """
    Create a small set of sample files inside test_files/ so there is always
    something to process on the first run, even without a real email inbox.
    Existing files are not overwritten.
    """
    samples = [
        ("invoice_april_2024.pdf",   b"%PDF-1.4 dummy invoice"),
        ("meeting_notes.docx",       b"PK dummy docx content"),
        ("budget_2024.xlsx",         b"PK dummy xlsx content"),
        ("project_plan.doc",         b"dummy doc content"),
        ("salary_sheet.xls",         b"dummy xls content"),
        ("team_photo.png",           b"\x89PNG dummy image"),
        ("product_screenshot.jpg",   b"\xff\xd8 dummy jpeg"),
        ("architecture_diagram.gif", b"GIF89a dummy gif"),
        ("readme_notes.txt",         b"This is a plain-text readme."),
        ("data_export.csv",          b"id,name,value\n1,Alice,100\n2,Bob,200"),
        ("presentation.pptx",        b"PK dummy pptx content"),
        ("unknown_format.dat",       b"arbitrary binary data"),
    ]

    created = 0
    for file_name, content in samples:
        dest = os.path.join(TEST_FILES_DIR, file_name)
        if not os.path.exists(dest):
            with open(dest, "wb") as fh:
                fh.write(content)
            created += 1

    if created:
        print(f"[✔] Created {created} sample file(s) in test_files/.")
    else:
        print("[ℹ] test_files/ already populated — skipping sample creation.")


def process_test_files() -> None:
    """Scan test_files/ and move every file to its category folder."""
    files = [
        f for f in os.listdir(TEST_FILES_DIR)
        if os.path.isfile(os.path.join(TEST_FILES_DIR, f))
    ]

    if not files:
        print("[ℹ] No files found in test_files/ to process.")
        return

    print(f"\n[→] Processing {len(files)} file(s) from test_files/...\n")
    for file_name in sorted(files):
        src = os.path.join(TEST_FILES_DIR, file_name)
        move_file_to_category(src, email_subject="Test Run", sender="local@test")


# ══════════════════════════════════════════════════════════════════════════════
# 7. MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    print("=" * 60)
    print("  Automated Email Attachment Organizer — Phase 2")
    print("=" * 60)

    print("\n[1] Setting up folders …")
    create_folders()

    print("\n[2] Setting up Excel report …")
    create_report_if_missing()

    print("\n[3] Creating sample test files …")
    create_sample_test_files()

    print("\n[4] Organising local test files …")
    process_test_files()

    print("\n[5] Downloading Gmail attachments …")
    total = download_gmail_attachments()
    if total > 0:
        print(f"\n[\u2714] Gmail: {total} attachment(s) downloaded and organised.")
    else:
        print("\n[\u2139] Gmail: 0 attachments processed "
              "(check .env credentials / keyword, or no matching emails).")

    print("\n[\u2714] Done!  Open reports/daily_report.xlsx to review the log.")
    print("=" * 60)


_INTERVAL_MINUTES = 30  # default; overridden by --interval


def gmail_job() -> None:
    """Lightweight job for scheduled runs — skips one-time setup."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n{chr(9472)*60}")
    print(f"  [\u23f0] Scheduled run @ {now}")
    print(chr(9472)*60)
    total = download_gmail_attachments()
    if total > 0:
        print(f"  [\u2714] {total} attachment(s) downloaded and organised.")
    else:
        print("  [\u2139] No new matching emails found.")
    print(f"  [\u2192] Next run in {_INTERVAL_MINUTES} minute(s). Press Ctrl+C to stop.\n")


if __name__ == "__main__":
    import argparse
    # pyrefly: ignore [missing-import]
    import schedule
    import time as _time

    parser = argparse.ArgumentParser(
        description="RPA Email Attachment Organizer",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 bot.py                            # run once (default)
  python3 bot.py --schedule                 # run every 30 minutes forever
  python3 bot.py --schedule --interval 60   # run every 60 minutes
        """,
    )
    parser.add_argument("--schedule", "-s", action="store_true",
                        help="Run continuously on a fixed interval (default: 30 min)")
    parser.add_argument("--interval", "-i", type=int, default=30, metavar="MINUTES",
                        help="Interval in minutes (default: 30). Requires --schedule.")
    args = parser.parse_args()
    _INTERVAL_MINUTES = args.interval

    if args.schedule:
        print("=" * 60)
        print(f"  RPA Bot — Scheduled Mode  (every {_INTERVAL_MINUTES} min)")
        print("  Press Ctrl+C to stop.")
        print("=" * 60)

        # One-time setup
        create_folders()
        create_report_if_missing()
        create_sample_test_files()
        process_test_files()

        # First Gmail fetch immediately
        gmail_job()

        # Register recurring job
        schedule.every(_INTERVAL_MINUTES).minutes.do(gmail_job)
        print(f"[\u2714] Scheduler running — next fetch in {_INTERVAL_MINUTES} min.\n")

        try:
            while True:
                schedule.run_pending()
                _time.sleep(15)
        except KeyboardInterrupt:
            print("\n[\u2139] Scheduler stopped. Goodbye!")
    else:
        main()
